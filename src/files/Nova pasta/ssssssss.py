import pyautogui
import openpyxl
import time
import sys

# --- Configurações ---
# Nome do arquivo Excel gerado no primeiro passo
ARQUIVO_EXCEL = 'produtos_nfe.xlsx'

# Coordenadas fornecidas pelo usuário
COORDENADAS = {
    "codigo": (509, 415),
    "qtd": (1025, 426),
    "vlr_unt": (1092, 424),
    "confirmacao": (1418, 426)
}

def automatizar_lancamento(arquivo_excel):
    """
    Lê o Excel e executa a sequência de cliques e digitação para cada item.
    """
    try:
        # Carrega o Workbook (livro)
        wb = openpyxl.load_workbook(arquivo_excel)
        ws = wb.active
    except FileNotFoundError:
        print(f"ERRO: Arquivo '{arquivo_excel}' não encontrado. Certifique-se de que ele foi gerado e está na mesma pasta.")
        sys.exit(1)
    except Exception as e:
        print(f"ERRO ao carregar o Excel: {e}")
        sys.exit(1)

    print("-" * 50)
    print("INICIANDO AUTOMAÇÃO...")
    print("Você tem 5 segundos para trocar para a janela do seu sistema.")
    print("-" * 50)
    
    # 5 SEGUNDOS DE ESPERA para o usuário posicionar o mouse e trocar de janela
    time.sleep(5) 
    
    # Itera sobre as linhas do Excel, começando da segunda (índice 2) para pular o cabeçalho
    # As colunas são: A (1) = Código, B (2) = Qtd, C (3) = Vlr Unitário
    itens_processados = 0
    
    # ws.iter_rows(min_row=2) começa a iteração a partir da segunda linha
    for row in ws.iter_rows(min_row=2):
        try:
            # Captura os dados da linha
            codigo = str(row[0].value) # Coluna A (Índice 0)
            qtd = str(row[1].value)    # Coluna B (Índice 1)
            vlr_unt = str(row[2].value) # Coluna C (Índice 2)

            print(f"Processando item: Código={codigo}, Qtd={qtd}, Vlr={vlr_unt}")
            
            # 1. CLICAR e DIGITAR o CÓDIGO (X=509, Y=415)
            pyautogui.click(COORDENADAS["codigo"])
            pyautogui.write(codigo)
            pyautogui.press('enter')

            # 2. CLICAR, APAGAR e DIGITAR a QUANTIDADE (X=912, Y=400)
            # Dois cliques para selecionar o conteúdo e BACKSPACE para garantir que apague
            #pyautogui.doubleClick(COORDENADAS["qtd"])
            pyautogui.press('backspace') # Garante que o campo esteja limpo
            pyautogui.write(qtd)
            pyautogui.press('enter')

            # 3. CLICAR, APAGAR e DIGITAR o VALOR UNITÁRIO (X=956, Y=411)
            #pyautogui.doubleClick(COORDENADAS["vlr_unt"])
            pyautogui.press('backspace') # Garante que o campo esteja limpo
            pyautogui.write(vlr_unt)
            pyautogui.press('enter')

            # 4. CLICAR no BOTÃO/CAMPO de CONFIRMAÇÃO (X=1423, Y=414)
            pyautogui.click(COORDENADAS["confirmacao"])
            
            # 5. ESPERAR 2 SEGUNDOS antes de ir para o próximo item
            time.sleep(2)
            
            itens_processados += 1
            
        except Exception as e:
            print(f"Ocorreu um erro ao processar a linha (código {codigo}): {e}")
            # Em caso de erro, você pode querer parar a automação ou apenas pular o item
            break 
            
    print("-" * 50)
    print(f"AUTOMAÇÃO CONCLUÍDA: {itens_processados} itens processados.")
    print("-" * 50)


if __name__ == "__main__":
    automatizar_lancamento(ARQUIVO_EXCEL)